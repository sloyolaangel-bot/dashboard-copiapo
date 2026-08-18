from pathlib import Path
from datetime import datetime, date
from zoneinfo import ZoneInfo
import json, re, unicodedata
from openpyxl import load_workbook

ROOT=Path(__file__).resolve().parents[1]
TZ=ZoneInfo("America/Santiago")
NOW=datetime.now(TZ)
EXCLUDED=("RUTH CABALLERO","MARIE CARMEN TORO","TORO ARANEDA")

def norm(v):
    s="" if v is None else str(v)
    s=unicodedata.normalize("NFD",s)
    s="".join(ch for ch in s if unicodedata.category(ch)!="Mn")
    return re.sub(r"\s+"," ",s).strip().upper()

def number(v):
    if v is None or v=="": return 0
    if isinstance(v,(int,float)): return float(v)
    try: return float(str(v).replace(",","."))
    except: return 0

MONTHS={
 "ENERO":1,"FEBRERO":2,"MARZO":3,"ABRIL":4,"MAYO":5,"JUNIO":6,
 "JULIO":7,"AGOSTO":8,"SEPTIEMBRE":9,"SETIEMBRE":9,"OCTUBRE":10,
 "NOVIEMBRE":11,"DICIEMBRE":12
}
MONTH_LABEL={v:k.lower() for k,v in MONTHS.items()}
MONTH_LABEL[9]="septiembre"

def parse_date_value(v):
    if isinstance(v,datetime): return v.date()
    if isinstance(v,date): return v
    s=str(v or "").strip()
    # dd/mm/yyyy
    m=re.search(r"(\d{1,2})[\/\-.](\d{1,2})[\/\-.](\d{2,4})",s)
    if m:
        y=int(m.group(3)); y=y+2000 if y<100 else y
        try:return date(y,int(m.group(2)),int(m.group(1)))
        except:return None
    # dd de agosto [de 2026]
    m=re.search(r"(\d{1,2})\s+DE\s+([A-ZÁÉÍÓÚÑ]+)(?:\s+DE\s+(\d{4}))?",norm(s))
    if m and norm(m.group(2)) in MONTHS:
        try:return date(int(m.group(3) or NOW.year),MONTHS[norm(m.group(2))],int(m.group(1)))
        except:return None
    # Common malformed report strings such as "Gestión al 10 de de"
    m=re.search(r"(?:GESTION|INFORMACION|DATOS|VENTAS|REPORTE)?\s*(?:AL|HASTA)\s*(?:EL\s*)?(\d{1,2})(?:\s+DE)?",norm(s))
    if m:
        day=int(m.group(1))
        if 1<=day<=31:
            try:return date(NOW.year,NOW.month,day)
            except:return None
    return None

def detect_report_date(wb, preferred_sheet=None):
    strong=[]
    weak=[]
    names=[]
    if preferred_sheet:
        names=[preferred_sheet]
    names += [n for n in wb.sheetnames if n not in names]
    label_words=("GESTION","ACTUALIZ","FECHA","DATOS","INFORMACION","REPORTE","VENTAS","CORTE")
    for sn in names:
        ws=wb[sn]
        # only first 45 rows and 80 cols; report header lives here, keeps Action quick
        for row in ws.iter_rows(min_row=1,max_row=min(ws.max_row,45),max_col=min(ws.max_column,80),values_only=True):
            for v in row:
                if v is None: continue
                d=parse_date_value(v)
                n=norm(v)
                if d:
                    if any(w in n for w in label_words): strong.append(d)
                    else: weak.append(d)
        if strong: break
    candidates=strong or weak
    # Ignore impossible old dates; current reports are 2026.
    candidates=[d for d in candidates if 2024<=d.year<=2035]
    return max(candidates) if candidates else date(NOW.year,NOW.month,NOW.day)

def report_meta(d):
    return {
      "report_date": d.isoformat(),
      "report_date_cl": d.strftime("%d/%m/%Y"),
      "report_label": f"Gestión al {d.day} de {MONTH_LABEL[d.month]}"
    }

def find_sheet(wb, wanted):
    nw=norm(wanted)
    for s in wb.sheetnames:
        if norm(s)==nw:return s
    for s in wb.sheetnames:
        if nw in norm(s):return s
    return None

def rows_values(ws):
    return list(ws.iter_rows(values_only=True))

def find_header(rows):
    for i,row in enumerate(rows):
        ns=[norm(x) for x in row]
        if "SUCURSAL" in ns and "NOMBRE" in ns:return i
    return -1

def col(headers,names,fallback):
    nh=[norm(x) for x in headers]
    for name in names:
        n=norm(name)
        if n in nh:return nh.index(n)
    return fallback

def is_excluded(name):
    n=norm(name)
    return any(x in n for x in EXCLUDED)

def write_json(path,obj):
    path.parent.mkdir(parents=True,exist_ok=True)
    path.write_text(json.dumps(obj,ensure_ascii=False,separators=(",",":")),encoding="utf-8")


def find_report_file(folder):
    folder = Path(folder)
    if not folder.exists():
        raise FileNotFoundError(f"No existe la carpeta: {folder}")
    preferred = folder / "reporte.xlsx"
    if preferred.exists():
        print(f"Usando archivo: {preferred}")
        return preferred
    candidates = sorted(
        [p for p in folder.glob("*.xlsx") if not p.name.startswith("~$")],
        key=lambda p: p.stat().st_mtime,
        reverse=True
    )
    if not candidates:
        raise FileNotFoundError(f"No se encontró ningún .xlsx dentro de {folder}")
    print(f"reporte.xlsx no encontrado. Usando el .xlsx más reciente: {candidates[0]}")
    return candidates[0]

def build_vida(path):
    path=Path(path)
    if not path.exists(): raise FileNotFoundError(f"Vida+Salud: no existe {path}")
    wb=load_workbook(path,read_only=True,data_only=True)
    sn=find_sheet(wb,"Detalle Ejec CCSS")
    if not sn: raise RuntimeError("Vida+Salud: no existe Detalle Ejec CCSS")
    rows=rows_values(wb[sn]); hr=find_header(rows)
    if hr<0: raise RuntimeError("Vida+Salud: encabezados no encontrados")
    h=rows[hr]
    cSuc=col(h,["Sucursal"],2); cPos=col(h,["Posición","Posicion"],4)
    cRut=col(h,["Rut"],5); cNom=col(h,["Nombre"],6)
    cVida=col(h,["Seg. Vida"],7); cSalud=col(h,["Seg. Salud"],9)
    data=[]
    for r in rows[hr+1:]:
        if cNom>=len(r):continue
        if norm(r[cSuc] if cSuc<len(r) else "")!="COPIAPO":continue
        if "ASISTENTE COMERCIAL" not in norm(r[cPos] if cPos<len(r) else ""):continue
        name=str(r[cNom] or "").strip()
        if not name or is_excluded(name):continue
        data.append({
          "rut":str(r[cRut] or "").strip() if cRut<len(r) else "",
          "nombre":name,
          "posicion":str(r[cPos] or "").strip(),
          "vida":number(r[cVida] if cVida<len(r) else 0),
          "salud":number(r[cSalud] if cSalud<len(r) else 0)
        })
    if not data: raise RuntimeError("Vida+Salud: no se encontraron asistentes comerciales de Copiapó")
    d=detect_report_date(wb,sn)
    print(f"Vida+Salud: {len(data)} asistentes encontrados. Fecha: {d}")
    obj={**report_meta(d),"generated_at":NOW.isoformat(),"data":data}
    write_json(ROOT/"vida-salud"/"reporte.json",obj)

    # Automatic weekly history. Store the latest cumulative snapshot for each campaign week.
    hp=ROOT/"vida-salud"/"history.json"
    try: history=json.loads(hp.read_text(encoding="utf-8")) if hp.exists() else {}
    except: history={}
    wk="S1" if d.day<=9 else "S2" if d.day<=16 else "S3" if d.day<=23 else "S4"
    history[wk]={"report_date":d.isoformat(),"report_label":obj["report_label"],"data":data}
    # Cierre oficial auditado de Semana 01–09 desde ZC 09082026.
    # Se fuerza este cierre para evitar que datos del 10-08 se asignen por error a S1.
    s1_official=[
      {"rut":"","nombre":"CAMILA CONSTANZA ROSALES CASTRO","posicion":"","vida":0,"salud":0},
      {"rut":"","nombre":"CATALINA ANAIS CASTRO CASTRO","posicion":"","vida":0,"salud":0},
      {"rut":"","nombre":"EDITH MAGALY TRONCOSO CASTILLO","posicion":"","vida":0,"salud":0},
      {"rut":"","nombre":"ESTER GLORIA URRUTIA GARIN","posicion":"","vida":10,"salud":0},
      {"rut":"","nombre":"EVA CRISTINA VEGA VILLARROEL","posicion":"","vida":16,"salud":0},
      {"rut":"","nombre":"MACARENA ALEJANDRA RIVERA GODOY","posicion":"","vida":0,"salud":0},
      {"rut":"","nombre":"MAYLIN DANIELA SOTO COLMAN","posicion":"","vida":9,"salud":0},
      {"rut":"","nombre":"MEY-GI SOLANGE LOCK CASTRO","posicion":"","vida":1,"salud":0},
      {"rut":"","nombre":"TERESA NYBROSKA CONTRERAS CARMONA","posicion":"","vida":14,"salud":0},
      {"rut":"","nombre":"VALERIA ANDREA PEREZ ROJAS","posicion":"","vida":0,"salud":0}
    ]
    history["S1"]={
      "report_date":f"{d.year}-{d.month:02d}-09",
      "report_label":f"Gestión al 9 de {MONTH_LABEL[d.month]}",
      "data":s1_official
    }
    write_json(hp,history)

def build_cruce(path):
    path=Path(path)
    if not path.exists(): raise FileNotFoundError(f"Cruce: no existe {path}")
    wb=load_workbook(path,read_only=True,data_only=True)
    sn=find_sheet(wb,"Detalle Ejec CCSS")
    if not sn: raise RuntimeError("Cruce: no existe Detalle Ejec CCSS")
    rows=rows_values(wb[sn]); hr=find_header(rows)
    h=rows[hr]
    cSuc=col(h,["Sucursal"],2);cPos=col(h,["Posición","Posicion"],4)
    cRut=col(h,["Rut"],5);cNom=col(h,["Nombre"],6)
    # Captación: source columns N:Q => zero-based 13:16
    data=[]
    for r in rows[hr+1:]:
        if len(r)<=16:continue
        if norm(r[cSuc])!="COPIAPO" or "ASISTENTE COMERCIAL" not in norm(r[cPos]):continue
        name=str(r[cNom] or "").strip()
        if not name or is_excluded(name):continue
        data.append({"rut":str(r[cRut] or "").strip(),"nombre":name,"posicion":str(r[cPos] or "").strip(),
          "segDesg":number(r[13]),"trxDesg":number(r[14]),"segCes":number(r[15]),"trxCes":number(r[16])})
    if not data: raise RuntimeError("Cruce: no se encontraron asistentes comerciales de Copiapó")
    d=detect_report_date(wb,sn)
    print(f"Cruce: {len(data)} asistentes encontrados. Fecha: {d}")
    write_json(ROOT/"cruce-captacion"/"reporte.json",{**report_meta(d),"generated_at":NOW.isoformat(),"data":data})

def build_avsav(path):
    path=Path(path)
    if not path.exists(): raise FileNotFoundError(f"AV-SAV: no existe {path}")
    wb=load_workbook(path,read_only=True,data_only=True)
    avsn=find_sheet(wb,"Av-Sav")
    dsn=find_sheet(wb,"Detalle Ejec CCSS")
    if not avsn or not dsn: raise RuntimeError("AV-SAV: faltan hojas requeridas")
    avrows=rows_values(wb[avsn])
    hr=-1
    for i,r in enumerate(avrows):
        if len(r)>4 and norm(r[1])=="COD_SUC_CCSS" and norm(r[4])=="SUCURSAL":hr=i;break
    if hr<0: raise RuntimeError("AV-SAV: tabla sucursal no encontrada")
    branchrow=next((r for r in avrows[hr+1:] if len(r)>4 and norm(r[4])=="COPIAPO"),None)
    if branchrow is None:raise RuntimeError("AV-SAV: Copiapó no encontrada")
    def rv(i):return number(branchrow[i]) if i<len(branchrow) else 0
    ofertaCes=rv(95);ofertaMixto=rv(97);segCes=rv(99);segMixto=rv(101)
    meta=rv(105) or .65; seguros=segCes+segMixto; ofertas=ofertaCes+ofertaMixto
    cruce=seguros/ofertas if ofertas else rv(107)
    branch={"seguros":seguros,"ofertas":ofertas,"cruce":cruce,"meta":meta,
            "cumplimiento":cruce/meta if meta else rv(109)}
    rows=rows_values(wb[dsn]); dhr=find_header(rows);h=rows[dhr]
    cSuc=col(h,["Sucursal"],2);cPos=col(h,["Posición","Posicion"],4)
    cNom=col(h,["Nombre"],6);cRut=col(h,["Rut"],5)
    cSegAv=col(h,["Seg. AV Ces"],22);cOfertaAv=col(h,["Trx Oferta Av"],25)
    cSegSav=col(h,["Seg. SAV Ces"],26);cOfertaSav=col(h,["Trx Oferta Sav"],29)
    detail=[]
    for r in rows[dhr+1:]:
        if cNom>=len(r):continue
        if norm(r[cSuc])!="COPIAPO" or "ASISTENTE COMERCIAL" not in norm(r[cPos]):continue
        name=str(r[cNom] or "").strip()
        if not name or is_excluded(name):continue
        def get(i):return r[i] if i<len(r) else 0
        detail.append({"sucursal":str(get(cSuc) or ""),"posicion":str(get(cPos) or ""),
          "rut":str(get(cRut) or ""),"nombre":name,"segAvCes":number(get(cSegAv)),
          "segSavCes":number(get(cSegSav)),"ofertaAv":number(get(cOfertaAv)),"ofertaSav":number(get(cOfertaSav))})
    if not detail: raise RuntimeError("AV-SAV: no se encontraron asistentes comerciales de Copiapó")
    d=detect_report_date(wb,dsn)
    print(f"AV-SAV: {len(detail)} asistentes encontrados. Fecha: {d}")
    write_json(ROOT/"av-sav-cesantia"/"reporte.json",{**report_meta(d),"generated_at":NOW.isoformat(),"branch":branch,"detail":detail})


def first_header(rows, required, max_rows=None):
    lim=len(rows) if max_rows is None else min(len(rows),max_rows)
    for i in range(lim):
        n=[norm(x) for x in rows[i]]
        if all(norm(req) in n for req in required):
            return i
    return -1

def find_col_range(headers,names,start=0,end=None):
    nh=[norm(x) for x in headers]
    end=len(nh) if end is None else min(end,len(nh))
    for name in names:
        nn=norm(name)
        for i in range(max(0,start),end):
            if nh[i]==nn:return i
    return -1

def build_gestion(path):
    path=Path(path)
    if not path.exists():raise FileNotFoundError(f"Gestión Comercial: no existe {path}")
    wb=load_workbook(path,read_only=True,data_only=True)

    gsn=find_sheet(wb,"Gestión Av + Sav") or find_sheet(wb,"Gestion Av + Sav")
    asn=find_sheet(wb,"Asistente Comercial")
    if not gsn or not asn:raise RuntimeError("Gestión Comercial: faltan hojas Gestión Av + Sav o Asistente Comercial")

    grows=rows_values(wb[gsn])
    gh=first_header(grows,["Sucursal","Venta","Meta"],30)
    if gh<0:raise RuntimeError("Gestión Comercial: encabezado Sucursal/Venta/Meta no encontrado")
    h=grows[gh]
    cSuc=find_col_range(h,["Sucursal"]);cVenta=find_col_range(h,["Venta"]);cMeta=find_col_range(h,["Meta"])
    cCum=find_col_range(h,["% Cump","%Cumpl","% Cumpl"])
    cCumA=find_col_range(h,["%Cump Acumulado","% Cump Acumulado","%Cumpl acumulado"])
    br=next((row for row in grows[gh+1:] if cSuc>=0 and cSuc<len(row) and norm(row[cSuc])=="COPIAPO"),None)
    if br is None:raise RuntimeError("Gestión Comercial: COPIAPO no encontrada en Gestión Av + Sav")
    def gv(row,i):return row[i] if i>=0 and i<len(row) else 0
    report=""
    for row in grows[:6]:
        for v in row:
            if "RESUMEN DEL" in norm(v):
                report=re.sub(r"^\s*Resumen del\s*","",str(v),flags=re.I).strip()
                break
        if report:break

    arows=rows_values(wb[asn])
    ah=first_header(arows,["Sucursal","Nombre","Av + Sav + Consumo"],15)
    if ah<0:raise RuntimeError("Gestión Comercial: encabezados de Asistente Comercial no encontrados")
    ha=arows[ah]
    aSuc=find_col_range(ha,["Sucursal"]);aNom=find_col_range(ha,["Nombre"]);aVenta=find_col_range(ha,["Av + Sav + Consumo"])
    secStart=-1;secEnd=len(ha)
    for sr in arows[:ah]:
        for i,v in enumerate(sr):
            if norm(v)=="GESTION CLIENTE (GESTION PROPIA)":
                secStart=i
                for j in range(i+1,len(sr)):
                    if norm(sr[j]):
                        secEnd=j;break
                break
        if secStart>=0:break
    if secStart<0:secStart,secEnd=60,69
    aCapt=find_col_range(ha,["Captación","Captacion"],secStart,secEnd)
    aCC=find_col_range(ha,["Cuenta Corriente"],secStart,secEnd)
    if min(aSuc,aNom,aVenta,aCapt,aCC)<0:
        raise RuntimeError("Gestión Comercial: faltan columnas AV+SAV+Consumo, Captación o Cuenta Corriente")
    assistants=[]
    for row in arows[ah+1:]:
        if max(aSuc,aNom,aVenta,aCapt,aCC)>=len(row):continue
        if norm(row[aSuc])!="COPIAPO":continue
        name=str(row[aNom] or "").strip()
        if not name:continue
        assistants.append({"nombre":name,"avSavConsumo":number(row[aVenta]),"captacion":number(row[aCapt]),"cuentaCorriente":number(row[aCC])})
    if not assistants:raise RuntimeError("Gestión Comercial: no se encontraron asistentes de Copiapó")
    obj={"generated_at":NOW.isoformat(),"reportDate":report or "Fecha no detectada",
         "branch":{"venta":number(gv(br,cVenta)),"meta":number(gv(br,cMeta)),"cumpl":number(gv(br,cCum)),"cumplAcum":number(gv(br,cCumA))},
         "assistants":assistants}
    print(f"Gestión Comercial: {len(assistants)} asistentes. Reporte: {obj['reportDate']}")
    write_json(ROOT/"gestion-comercial"/"reporte.json",obj)

def date_text_py(v):
    if isinstance(v,datetime):return v.strftime("%d/%m/%Y")
    if isinstance(v,date):return v.strftime("%d/%m/%Y")
    s=str(v or "").strip()
    m=re.search(r"(\d{4})[-/](\d{1,2})[-/](\d{1,2})",s)
    if m:return f"{int(m.group(3)):02d}/{int(m.group(2)):02d}/{m.group(1)}"
    return s

def build_renegociacion(path):
    path=Path(path)
    if not path.exists():raise FileNotFoundError(f"Renegociación: no existe {path}")
    wb=load_workbook(path,read_only=True,data_only=True)
    rsn=find_sheet(wb,"Resumen General");dsn=find_sheet(wb,"Detalle AACC")
    if not rsn or not dsn:raise RuntimeError("Renegociación: faltan hojas Resumen General o Detalle AACC")
    rr=rows_values(wb[rsn])
    rh=first_header(rr,["Sucursal","Meta Rene Mm$","Q Total Rene","Monto Rene Mm$"])
    if rh<0:rh=first_header(rr,["Sucursal","Q Total Rene"])
    if rh<0:raise RuntimeError("Renegociación: encabezado de Resumen General no encontrado")
    hr=rr[rh]
    cSuc=find_col_range(hr,["Sucursal"]);cMeta=find_col_range(hr,["Meta Rene Mm$","Meta Rene"])
    cQ=find_col_range(hr,["Q Total Rene"]);cMonto=find_col_range(hr,["Monto Rene Mm$","Monto Rene"])
    cCum=find_col_range(hr,["%Cump. Rene","% Cump. Rene","Cump.Rene"])
    br=next((row for row in rr[rh+1:] if cSuc>=0 and cSuc<len(row) and norm(row[cSuc])=="COPIAPO"),None)
    if br is None:raise RuntimeError("Renegociación: COPIAPO no encontrada en Resumen General")
    def rv(row,i):return row[i] if i>=0 and i<len(row) else 0
    fecha=""
    for row in rr:
        for p,v in enumerate(row):
            if norm(v)=="ACTUALIZACION RENEGOCIACIONES":
                fecha=date_text_py(row[p+1] if p+1<len(row) else "");break
        if fecha:break
    if not fecha:
        ausn=find_sheet(wb,"Actualizacion") or find_sheet(wb,"Actualización")
        if ausn:
            ar=rows_values(wb[ausn])
            row=next((x for x in ar if len(x)>1 and norm(x[0])=="RENEGOCIACIONES"),None)
            if row:fecha=date_text_py(row[1])

    dr=rows_values(wb[dsn])
    dh=first_header(dr,["Nombre Aacc","Sucursal","Meta Rene","Q Total Rene"])
    if dh<0:raise RuntimeError("Renegociación: encabezado Detalle AACC no encontrado")
    hd=dr[dh]
    dNom=find_col_range(hd,["Nombre Aacc"]);dSuc=find_col_range(hd,["Sucursal"])
    dMeta=find_col_range(hd,["Meta Rene"]);dQ=find_col_range(hd,["Q Total Rene"])
    dMonto=find_col_range(hd,["Monto Rene Mm$","Monto Rene"]);dCum=find_col_range(hd,["%Cump. Rene","Cump.Rene"])
    detail=[]
    for row in dr[dh+1:]:
        if max(dNom,dSuc,dMeta,dQ,dMonto,dCum)>=len(row):continue
        if norm(row[dSuc])!="COPIAPO":continue
        name=str(row[dNom] or "").strip()
        if not name:continue
        detail.append({"nombre":name,"meta":number(row[dMeta]),"q":number(row[dQ]),"monto":number(row[dMonto]),"cumpl":number(row[dCum])})
    if not detail:raise RuntimeError("Renegociación: no se encontraron asistentes de Copiapó")
    branch={"fecha":fecha or "—","meta":number(rv(br,cMeta)),"q":number(rv(br,cQ)),"monto":number(rv(br,cMonto)),"cumpl":number(rv(br,cCum))}
    print(f"Renegociación: {len(detail)} asistentes. Fecha: {branch['fecha']}")
    write_json(ROOT/"renegociacion"/"reporte.json",{"generated_at":NOW.isoformat(),"branch":branch,"detail":detail})

if __name__=="__main__":
    print(f"ROOT detectado: {ROOT}")
    vida_file=find_report_file(ROOT/"vida-salud")
    cruce_file=find_report_file(ROOT/"cruce-captacion")
    avsav_file=find_report_file(ROOT/"av-sav-cesantia")
    gestion_file=find_report_file(ROOT/"gestion-comercial")
    reneg_file=find_report_file(ROOT/"renegociacion")

    build_vida(vida_file)
    build_cruce(cruce_file)
    build_avsav(avsav_file)
    build_gestion(gestion_file)
    build_renegociacion(reneg_file)

    print("JSON de los 5 dashboards generados correctamente.")
