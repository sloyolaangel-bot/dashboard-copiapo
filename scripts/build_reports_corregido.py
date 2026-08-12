from pathlib import Path
from datetime import datetime, date
from zoneinfo import ZoneInfo
import json, re, unicodedata
from openpyxl import load_workbook

ROOT=Path(__file__).resolve().parents[1]
TZ=ZoneInfo("America/Santiago")
NOW=datetime.now(TZ)
EXCLUDED=("RUTH CABALLERO","MARIE CARMEN TORO","TORO ARANEDA")

def norm(v):
    s="" if v is None else str(v)
    s=unicodedata.normalize("NFD",s)
    s="".join(ch for ch in s if unicodedata.category(ch)!="Mn")
    return re.sub(r"\s+"," ",s).strip().upper()

def number(v):
    if v is None or v=="": return 0
    if isinstance(v,(int,float)): return float(v)
    try: return float(str(v).replace(",","."))
    except: return 0

MONTHS={
 "ENERO":1,"FEBRERO":2,"MARZO":3,"ABRIL":4,"MAYO":5,"JUNIO":6,
 "JULIO":7,"AGOSTO":8,"SEPTIEMBRE":9,"SETIEMBRE":9,"OCTUBRE":10,
 "NOVIEMBRE":11,"DICIEMBRE":12
}
MONTH_LABEL={v:k.lower() for k,v in MONTHS.items()}
MONTH_LABEL[9]="septiembre"

def parse_date_value(v):
    if isinstance(v,datetime): return v.date()
    if isinstance(v,date): return v
    s=str(v or "").strip()
    # dd/mm/yyyy
    m=re.search(r"(\d{1,2})[\/\-.](\d{1,2})[\/\-.](\d{2,4})",s)
    if m:
        y=int(m.group(3)); y=y+2000 if y<100 else y
        try:return date(y,int(m.group(2)),int(m.group(1)))
        except:return None
    # dd de agosto [de 2026]
    m=re.search(r"(\d{1,2})\s+DE\s+([A-ZÁÉÍÓÚÑ]+)(?:\s+DE\s+(\d{4}))?",norm(s))
    if m and norm(m.group(2)) in MONTHS:
        try:return date(int(m.group(3) or NOW.year),MONTHS[norm(m.group(2))],int(m.group(1)))
        except:return None
    # Common malformed report strings such as "Gestión al 10 de de"
    m=re.search(r"(?:GESTION|INFORMACION|DATOS|VENTAS|REPORTE)?\s*(?:AL|HASTA)\s*(?:EL\s*)?(\d{1,2})(?:\s+DE)?",norm(s))
    if m:
        day=int(m.group(1))
        if 1<=day<=31:
            try:return date(NOW.year,NOW.month,day)
            except:return None
    return None

def detect_report_date(wb, preferred_sheet=None):
    strong=[]
    weak=[]
    names=[]
    if preferred_sheet:
        names=[preferred_sheet]
    names += [n for n in wb.sheetnames if n not in names]
    label_words=("GESTION","ACTUALIZ","FECHA","DATOS","INFORMACION","REPORTE","VENTAS","CORTE")
    for sn in names:
        ws=wb[sn]
        # only first 45 rows and 80 cols; report header lives here, keeps Action quick
        for row in ws.iter_rows(min_row=1,max_row=min(ws.max_row,45),max_col=min(ws.max_column,80),values_only=True):
            for v in row:
                if v is None: continue
                d=parse_date_value(v)
                n=norm(v)
                if d:
                    if any(w in n for w in label_words): strong.append(d)
                    else: weak.append(d)
        if strong: break
    candidates=strong or weak
    # Ignore impossible old dates; current reports are 2026.
    candidates=[d for d in candidates if 2024<=d.year<=2035]
    return max(candidates) if candidates else date(NOW.year,NOW.month,NOW.day)

def report_meta(d):
    return {
      "report_date": d.isoformat(),
      "report_date_cl": d.strftime("%d/%m/%Y"),
      "report_label": f"Gestión al {d.day} de {MONTH_LABEL[d.month]}"
    }

def find_sheet(wb, wanted):
    nw=norm(wanted)
    for s in wb.sheetnames:
        if norm(s)==nw:return s
    for s in wb.sheetnames:
        if nw in norm(s):return s
    return None

def rows_values(ws):
    return list(ws.iter_rows(values_only=True))

def find_header(rows):
    for i,row in enumerate(rows):
        ns=[norm(x) for x in row]
        if "SUCURSAL" in ns and "NOMBRE" in ns:return i
    return -1

def col(headers,names,fallback):
    nh=[norm(x) for x in headers]
    for name in names:
        n=norm(name)
        if n in nh:return nh.index(n)
    return fallback

def is_excluded(name):
    n=norm(name)
    return any(x in n for x in EXCLUDED)

def write_json(path,obj):
    path.parent.mkdir(parents=True,exist_ok=True)
    path.write_text(json.dumps(obj,ensure_ascii=False,separators=(",",":")),encoding="utf-8")


def find_report_file(folder):
    folder = Path(folder)
    if not folder.exists():
        raise FileNotFoundError(f"No existe la carpeta: {folder}")
    preferred = folder / "reporte.xlsx"
    if preferred.exists():
        print(f"Usando archivo: {preferred}")
        return preferred
    candidates = sorted(
        [p for p in folder.glob("*.xlsx") if not p.name.startswith("~$")],
        key=lambda p: p.stat().st_mtime,
        reverse=True
    )
    if not candidates:
        raise FileNotFoundError(f"No se encontró ningún .xlsx dentro de {folder}")
    print(f"reporte.xlsx no encontrado. Usando el .xlsx más reciente: {candidates[0]}")
    return candidates[0]

def build_vida(path):
    path=Path(path)
    if not path.exists(): raise FileNotFoundError(f"Vida+Salud: no existe {path}")
    wb=load_workbook(path,read_only=True,data_only=True)
    sn=find_sheet(wb,"Detalle Ejec CCSS")
    if not sn: raise RuntimeError("Vida+Salud: no existe Detalle Ejec CCSS")
    rows=rows_values(wb[sn]); hr=find_header(rows)
    if hr<0: raise RuntimeError("Vida+Salud: encabezados no encontrados")
    h=rows[hr]
    cSuc=col(h,["Sucursal"],2); cPos=col(h,["Posición","Posicion"],4)
    cRut=col(h,["Rut"],5); cNom=col(h,["Nombre"],6)
    cVida=col(h,["Seg. Vida"],7); cSalud=col(h,["Seg. Salud"],9)
    data=[]
    for r in rows[hr+1:]:
        if cNom>=len(r):continue
        if norm(r[cSuc] if cSuc<len(r) else "")!="COPIAPO":continue
        if "ASISTENTE COMERCIAL" not in norm(r[cPos] if cPos<len(r) else ""):continue
        name=str(r[cNom] or "").strip()
        if not name or is_excluded(name):continue
        data.append({
          "rut":str(r[cRut] or "").strip() if cRut<len(r) else "",
          "nombre":name,
          "posicion":str(r[cPos] or "").strip(),
          "vida":number(r[cVida] if cVida<len(r) else 0),
          "salud":number(r[cSalud] if cSalud<len(r) else 0)
        })
    if not data: raise RuntimeError("Vida+Salud: no se encontraron asistentes comerciales de Copiapó")
    d=detect_report_date(wb,sn)
    print(f"Vida+Salud: {len(data)} asistentes encontrados. Fecha: {d}")
    obj={**report_meta(d),"generated_at":NOW.isoformat(),"data":data}
    write_json(ROOT/"vida-salud"/"reporte.json",obj)

    # Automatic weekly history. Store the latest cumulative snapshot for each campaign week.
    hp=ROOT/"vida-salud"/"history.json"
    try: history=json.loads(hp.read_text(encoding="utf-8")) if hp.exists() else {}
    except: history={}
    wk="S1" if d.day<=9 else "S2" if d.day<=16 else "S3" if d.day<=23 else "S4"
    history[wk]={"report_date":d.isoformat(),"report_label":obj["report_label"],"data":data}
    # First-week close previously established in the dashboard; preserve it if no S1 exists yet.
    if "S1" not in history:
        history["S1"]={"report_date":f"{d.year}-{d.month:02d}-09","report_label":f"Gestión al 9 de {MONTH_LABEL[d.month]}","data":[
          {"rut":"","nombre":"EVA CRISTINA VEGA VILLARROEL","posicion":"","vida":20,"salud":0},
          {"rut":"","nombre":"TERESA NYBROSKA CONTRERAS CARMONA","posicion":"","vida":14,"salud":0},
          {"rut":"","nombre":"ESTER GLORIA URRUTIA GARIN","posicion":"","vida":12,"salud":0},
          {"rut":"","nombre":"MAYLIN DANIELA SOTO COLMAN","posicion":"","vida":10,"salud":0},
          {"rut":"","nombre":"MEY-GI SOLANGE LOCK CASTRO","posicion":"","vida":1,"salud":0},
          {"rut":"","nombre":"CATALINA ANAIS CASTRO CASTRO","posicion":"","vida":1,"salud":0},
          {"rut":"","nombre":"MACARENA ALEJANDRA RIVERA GODOY","posicion":"","vida":1,"salud":0},
          {"rut":"","nombre":"CAMILA CONSTANZA ROSALES CASTRO","posicion":"","vida":0,"salud":0},
          {"rut":"","nombre":"VALERIA ANDREA PEREZ ROJAS","posicion":"","vida":0,"salud":0},
          {"rut":"","nombre":"EDITH MAGALY TRONCOSO CASTILLO","posicion":"","vida":0,"salud":0}
        ]}
    write_json(hp,history)

def build_cruce(path):
    path=Path(path)
    if not path.exists(): raise FileNotFoundError(f"Cruce: no existe {path}")
    wb=load_workbook(path,read_only=True,data_only=True)
    sn=find_sheet(wb,"Detalle Ejec CCSS")
    if not sn: raise RuntimeError("Cruce: no existe Detalle Ejec CCSS")
    rows=rows_values(wb[sn]); hr=find_header(rows)
    h=rows[hr]
    cSuc=col(h,["Sucursal"],2);cPos=col(h,["Posición","Posicion"],4)
    cRut=col(h,["Rut"],5);cNom=col(h,["Nombre"],6)
    # Captación: source columns N:Q => zero-based 13:16
    data=[]
    for r in rows[hr+1:]:
        if len(r)<=16:continue
        if norm(r[cSuc])!="COPIAPO" or "ASISTENTE COMERCIAL" not in norm(r[cPos]):continue
        name=str(r[cNom] or "").strip()
        if not name or is_excluded(name):continue
        data.append({"rut":str(r[cRut] or "").strip(),"nombre":name,"posicion":str(r[cPos] or "").strip(),
          "segDesg":number(r[13]),"trxDesg":number(r[14]),"segCes":number(r[15]),"trxCes":number(r[16])})
    if not data: raise RuntimeError("Cruce: no se encontraron asistentes comerciales de Copiapó")
    d=detect_report_date(wb,sn)
    print(f"Cruce: {len(data)} asistentes encontrados. Fecha: {d}")
    write_json(ROOT/"cruce-captacion"/"reporte.json",{**report_meta(d),"generated_at":NOW.isoformat(),"data":data})

def build_avsav(path):
    path=Path(path)
    if not path.exists(): raise FileNotFoundError(f"AV-SAV: no existe {path}")
    wb=load_workbook(path,read_only=True,data_only=True)
    avsn=find_sheet(wb,"Av-Sav")
    dsn=find_sheet(wb,"Detalle Ejec CCSS")
    if not avsn or not dsn: raise RuntimeError("AV-SAV: faltan hojas requeridas")
    avrows=rows_values(wb[avsn])
    hr=-1
    for i,r in enumerate(avrows):
        if len(r)>4 and norm(r[1])=="COD_SUC_CCSS" and norm(r[4])=="SUCURSAL":hr=i;break
    if hr<0: raise RuntimeError("AV-SAV: tabla sucursal no encontrada")
    branchrow=next((r for r in avrows[hr+1:] if len(r)>4 and norm(r[4])=="COPIAPO"),None)
    if branchrow is None:raise RuntimeError("AV-SAV: Copiapó no encontrada")
    def rv(i):return number(branchrow[i]) if i<len(branchrow) else 0
    ofertaCes=rv(95);ofertaMixto=rv(97);segCes=rv(99);segMixto=rv(101)
    meta=rv(105) or .65; seguros=segCes+segMixto; ofertas=ofertaCes+ofertaMixto
    cruce=seguros/ofertas if ofertas else rv(107)
    branch={"seguros":seguros,"ofertas":ofertas,"cruce":cruce,"meta":meta,
            "cumplimiento":cruce/meta if meta else rv(109)}
    rows=rows_values(wb[dsn]); dhr=find_header(rows);h=rows[dhr]
    cSuc=col(h,["Sucursal"],2);cPos=col(h,["Posición","Posicion"],4)
    cNom=col(h,["Nombre"],6);cRut=col(h,["Rut"],5)
    cSegAv=col(h,["Seg. AV Ces"],22);cOfertaAv=col(h,["Trx Oferta Av"],25)
    cSegSav=col(h,["Seg. SAV Ces"],26);cOfertaSav=col(h,["Trx Oferta Sav"],29)
    detail=[]
    for r in rows[dhr+1:]:
        if cNom>=len(r):continue
        if norm(r[cSuc])!="COPIAPO" or "ASISTENTE COMERCIAL" not in norm(r[cPos]):continue
        name=str(r[cNom] or "").strip()
        if not name or is_excluded(name):continue
        def get(i):return r[i] if i<len(r) else 0
        detail.append({"sucursal":str(get(cSuc) or ""),"posicion":str(get(cPos) or ""),
          "rut":str(get(cRut) or ""),"nombre":name,"segAvCes":number(get(cSegAv)),
          "segSavCes":number(get(cSegSav)),"ofertaAv":number(get(cOfertaAv)),"ofertaSav":number(get(cOfertaSav))})
    if not detail: raise RuntimeError("AV-SAV: no se encontraron asistentes comerciales de Copiapó")
    d=detect_report_date(wb,dsn)
    print(f"AV-SAV: {len(detail)} asistentes encontrados. Fecha: {d}")
    write_json(ROOT/"av-sav-cesantia"/"reporte.json",{**report_meta(d),"generated_at":NOW.isoformat(),"branch":branch,"detail":detail})

if __name__=="__main__":
    print(f"ROOT detectado: {ROOT}")
    vida_file=find_report_file(ROOT/"vida-salud")
    cruce_file=find_report_file(ROOT/"cruce-captacion")
    avsav_file=find_report_file(ROOT/"av-sav-cesantia")

    build_vida(vida_file)
    build_cruce(cruce_file)
    build_avsav(avsav_file)

    print("JSON generados correctamente.")
